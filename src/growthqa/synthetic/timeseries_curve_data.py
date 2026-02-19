#!/usr/bin/env python3
"""
timeseries_curve_data.py
---------------------------------
Synthetic growth-curve generator that writes a single **wide** CSV:

  timeseries_wide_<file-stem>.csv
    Columns:
      FileName, Test Id, Model Name, Is_Valid,
      T1 = 0 (<unit>), T2 = <t2> (<unit>), ...

Also writes run_info.xlsx exactly as before.

Notes:
  - Models: Logistic, Gompertz, ModifiedGompertz, Richards, Diauxic, Flat
  - Optional missing values & negative outlier injection
"""

import argparse
import logging
import os
import re
from datetime import datetime
import numpy as np
import pandas as pd

# --------------------------
# 1) Model definitions
# --------------------------
def logistic(t, A, mu, lam):
    return A / (1.0 + np.exp((4.0 * mu / A) * (lam - t) + 2.0))

def gompertz(t, A, mu, lam):
    return A * np.exp(-np.exp((mu * np.e / A) * (lam - t) + 1.0))

def modified_gompertz(t, A, mu, lam, alpha, tshift):
    part1 = A * np.exp(-np.exp((mu * np.e / A) * (lam - t) + 1.0))
    part2 = A * np.exp(alpha * (t - tshift))
    return part1 + part2

def richards(t, A, mu, lam, nu):
    return A * (1.0 + nu * np.exp((mu * (1.0 + nu) / A) * (lam - t))) ** (-1.0 / nu)

def diauxic(t, A1, mu1, lam1, A2, mu2, lam2):
    return logistic(t, A1, mu1, lam1) + logistic(t, A2, mu2, lam2)

def flat_line(t, baseline):
    return np.full_like(t, baseline)

# model name -> (fn, param names)
MODEL_SPECS = {
    "Logistic": (logistic, ["A", "mu", "lam"]),
    "Gompertz": (gompertz, ["A", "mu", "lam"]),
    "ModifiedGompertz": (modified_gompertz, ["A", "mu", "lam", "alpha", "tshift"]),
    "Richards": (richards, ["A", "mu", "lam", "nu"]),
    "Diauxic": (diauxic, ["A1", "mu1", "lam1", "A2", "mu2", "lam2"]),
    "Flat": (flat_line, ["baseline"]),
}

GOOD_MODELS = {"Logistic", "Gompertz", "ModifiedGompertz", "Richards"}

SUPPORTED_HORIZONS = [8.0, 14.75, 16.0, 24.0, 32.0, 52.0]
SUPPORTED_STEPS = [0.25, 0.5, 1.0]
SUPPORTED_BUCKETS = [(h, s) for h in SUPPORTED_HORIZONS for s in SUPPORTED_STEPS]

MIN_SYNTHETIC_PER_BUCKET = 800
SYNTHETIC_LAB_RATIO = 10
MIN_PER_SUBTYPE = 50
MAX_SYNTHETIC_PER_BUCKET = 3000

VALID_SUBTYPES = ["plain", "fast", "late", "diauxic", "decline"]
INVALID_SUBTYPES = ["nearreal", "subtle", "obvious", "decline_only", "oscillation", "noise", "collapse"]

# --------------------------
# 2) Noise / corruption
# --------------------------
def inject_missing(y: pd.Series, frac: float, rng: np.random.Generator):
    if frac <= 0: return y
    n = len(y)
    k = max(1, int(round(frac * n)))
    idx = rng.choice(n, size=min(k, n), replace=False)
    y.iloc[idx] = np.nan
    return y

def inject_negative_outliers(y: pd.Series, frac: float, scale_min: float, scale_max: float, rng: np.random.Generator):
    if frac <= 0: return y
    n = len(y)
    k = max(1, int(round(frac * n)))
    idx = rng.choice(n, size=min(k, n), replace=False)
    sub = rng.uniform(scale_min, scale_max, size=len(idx))
    y.iloc[idx] = (y.iloc[idx].values - sub).clip(min=0.0)
    return y

def make_obvious_invalid(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Very obvious invalid curve:
    - big drop to near-zero in the middle or tail.
    """
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 4:
        return y

    # Choose a cut in middle third
    cut = rng.integers(n // 3, 2 * n // 3)
    drop_factor = rng.uniform(0.0, 0.2)  # drop to 0–20% of value
    y[cut:] = y[cut] * drop_factor
    return y


def make_subtle_invalid(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Subtle invalid curve:
    - local dip segment (e.g. bubble / misreading) but not totally crazy.
    """
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 6:
        return y

    center = rng.integers(n // 4, 3 * n // 4)
    width = rng.integers(2, min(6, n - center))
    max_y = np.nanmax(y) if np.any(np.isfinite(y)) else 1.0
    drop = rng.uniform(0.1, 0.4) * max_y
    y[center:center + width] = np.clip(y[center:center + width] - drop, 0.0, None)
    return y


def make_near_real_invalid(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Very close to realistic but technically invalid curve:
    - tail is suppressed / slightly declining so it never really stabilises.
    """
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 6:
        return y

    start_tail = int(0.6 * n)
    factor = rng.uniform(0.4, 0.8)  # reduce plateau
    y[start_tail:] = y[start_tail:] * factor

    # optional very mild downward trend
    trend = np.linspace(0.0, rng.uniform(0.05, 0.15) * np.nanmax(y), n - start_tail)
    y[start_tail:] = np.clip(y[start_tail:] - trend, 0.0, None)
    return y

def make_decline_only_invalid(t: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Invalid: starts high and declines without a preceding rise.
    """
    t = np.asarray(t, dtype=float)
    baseline = rng.uniform(0.4, 1.2)
    k = rng.uniform(0.05, 0.2)
    y = baseline * np.exp(-k * (t - t.min()))
    return y

def make_chaotic_oscillation(t: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Invalid: high-frequency oscillations without clear structure.
    """
    t = np.asarray(t, dtype=float)
    base = rng.uniform(0.1, 0.4)
    amp = rng.uniform(0.2, 0.6)
    freq = rng.uniform(2.0, 6.0)
    y = base + amp * np.sin(freq * t + rng.uniform(0, 2 * np.pi))
    y += rng.normal(0, 0.15, size=len(t))
    return np.clip(y, 0.0, None)

def make_noise_dominated(t: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Invalid: noise-dominated, low signal-to-noise.
    """
    t = np.asarray(t, dtype=float)
    baseline = rng.uniform(0.05, 0.2)
    y = baseline + rng.normal(0, 0.2, size=len(t))
    return np.clip(y, 0.0, None)

def make_sudden_collapse(y: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    """
    Invalid: abrupt collapse in tail (more abrupt than valid decline).
    """
    y = np.asarray(y, dtype=float).copy()
    n = len(y)
    if n < 4:
        return y
    cut = rng.integers(n // 2, n - 1)
    y[cut:] = y[cut:] * rng.uniform(0.05, 0.25)
    return y

def apply_valid_decline(y: np.ndarray, t: np.ndarray, rng: np.random.Generator, t_start: float, k: float) -> np.ndarray:
    """
    Valid: structured decline after a clear growth phase.
    """
    y = np.asarray(y, dtype=float).copy()
    decay = np.ones_like(t, dtype=float)
    mask = t > t_start
    decay[mask] = np.exp(-k * (t[mask] - t_start))
    y = y * decay
    return np.clip(y, 0.0, None)

def _allocate_counts(total: int, weights: dict, min_each: int) -> dict:
    keys = list(weights.keys())
    if total <= 0:
        return {k: 0 for k in keys}
    if min_each * len(keys) > total:
        min_each = total // len(keys)
    counts = {k: min_each for k in keys}
    remaining = total - min_each * len(keys)
    if remaining <= 0:
        # distribute any leftover one-by-one if needed
        for i in range(total - sum(counts.values())):
            counts[keys[i % len(keys)]] += 1
        return counts
    w = np.array([max(0.0, weights[k]) for k in keys], dtype=float)
    if w.sum() <= 0:
        w = np.ones(len(keys), dtype=float)
    w = w / w.sum()
    extra = np.floor(w * remaining).astype(int)
    for k, v in zip(keys, extra):
        counts[k] += int(v)
    leftover = total - sum(counts.values())
    if leftover > 0:
        order = np.argsort(-w)
        for i in range(leftover):
            counts[keys[order[i % len(keys)]]] += 1
    return counts


def _safe_profile_from_filename(path: str) -> tuple[float, float] | None:
    name = os.path.basename(str(path))
    m = re.search(r"_(\d+(?:\.\d+)?)h_(\d+(?:\.\d+)?)\.csv$", name, flags=re.I)
    if not m:
        return None
    return float(m.group(1)), float(m.group(2))


def _safe_profile_from_columns(path: str) -> tuple[float, float] | None:
    try:
        df = pd.read_csv(path, nrows=1, low_memory=False)
    except Exception:
        return None
    times = []
    for c in df.columns:
        m = re.match(r"^T\s*([0-9]*\.?[0-9]+)\s*\(", str(c))
        if m:
            times.append(float(m.group(1)))
    if len(times) < 2:
        return None
    times = np.array(sorted(set(times)), dtype=float)
    diffs = np.diff(times)
    diffs = diffs[diffs > 0]
    if len(diffs) == 0:
        return None
    step = float(np.round(np.median(diffs), 3))
    tmax = float(np.round(np.max(times), 3))
    return tmax, step


def _nearest_supported_bucket(tmax: float, step: float) -> tuple[float, float] | None:
    best = None
    best_score = float("inf")
    for h, s in SUPPORTED_BUCKETS:
        score = abs(h - tmax) * 100 + abs(s - step) * 1000
        if score < best_score:
            best_score = score
            best = (h, s)
    if best is None:
        return None
    if abs(best[0] - tmax) <= 0.6 and abs(best[1] - step) <= 0.15:
        return best
    return None


def detect_lab_bucket_map(train_data_dir: str | os.PathLike) -> dict[tuple[float, float], list[str]]:
    train_data_dir = os.fspath(train_data_dir)
    mapping = {b: [] for b in SUPPORTED_BUCKETS}
    excluded = {"meta.csv", "raw_merged.csv", "final_merged.csv", "syn.csv"}
    for fn in os.listdir(train_data_dir):
        if not fn.lower().endswith(".csv"):
            continue
        if fn in excluded:
            continue
        path = os.path.join(train_data_dir, fn)
        prof = _safe_profile_from_filename(path) or _safe_profile_from_columns(path)
        if prof is None:
            continue
        bucket = _nearest_supported_bucket(*prof)
        if bucket is not None:
            mapping[bucket].append(path)
    return mapping


def _sample_good_model_for_subtype(
    *,
    subtype: str,
    time_points: np.ndarray,
    max_time: float,
    rng: np.random.Generator,
) -> tuple[str, np.ndarray]:
    model_name = "Diauxic" if subtype == "diauxic" else rng.choice(sorted(GOOD_MODELS))
    if model_name == "Diauxic":
        pars = {
            "A1": rng.uniform(0.3, 1.0),
            "mu1": rng.uniform(0.2, 1.2),
            "lam1": rng.uniform(0.0, 5.0),
            "A2": rng.uniform(0.3, 1.0),
            "mu2": rng.uniform(0.2, 1.2),
            "lam2": rng.uniform(5.0, min(15.0, max_time)),
        }
        return model_name, diauxic(time_points, **pars)

    late_lam_range = (8.0, 12.0) if max_time <= 16.0 else (10.0, min(16.0, max_time))
    if subtype == "fast":
        mu_range = (0.9, 1.5)
        lam_range = (0.0, 2.0)
    elif subtype == "late":
        mu_range = (0.8, 1.5)
        lam_range = late_lam_range
    else:
        mu_range = (0.2, 1.5)
        lam_range = (0.0, min(10.0, max_time))

    pars = {"A": rng.uniform(0.5, 2.0), "mu": rng.uniform(*mu_range), "lam": rng.uniform(*lam_range)}
    if model_name == "ModifiedGompertz":
        pars.update({"alpha": rng.uniform(0.0, 0.3), "tshift": rng.uniform(5.0, min(12.0, max_time))})
    if model_name == "Richards":
        pars.update({"nu": rng.uniform(0.5, 2.0)})
    return model_name, MODEL_SPECS[model_name][0](time_points, **pars)


def generate_bucket_synthetic_df(
    *,
    bucket_h: float,
    bucket_step: float,
    total_curves: int,
    file_stem: str,
    seed: int = 123,
    time_unit: str = "h",
    noise_level: float = 0.05,
    pct_high_quality_valid: float = 0.3,
    pct_missing_curves: float = 0.1,
    missing_frac_per_curve: float = 0.1,
    pct_outlier_curves: float = 0.05,
    outlier_frac_per_curve: float = 0.05,
    outlier_scale_min: float = 0.1,
    outlier_scale_max: float = 0.3,
    pct_valid_decline: float = 0.15,
    pct_fast_growth: float = 0.15,
    pct_late_growth: float = 0.15,
    valid_decline_k_min: float = 0.03,
    valid_decline_k_max: float = 0.15,
    pct_obvious_invalid_curves: float = 0.1,
    pct_subtle_invalid_curves: float = 0.1,
    pct_nearreal_invalid_curves: float = 0.05,
    pct_invalid_decline_only: float = 0.08,
    pct_invalid_oscillation: float = 0.06,
    pct_invalid_noise: float = 0.06,
    pct_invalid_collapse: float = 0.08,
) -> tuple[pd.DataFrame, dict]:
    rng = np.random.default_rng(seed)
    time_points = np.arange(0, bucket_h + bucket_step / 2, bucket_step)

    valid_target = int(round(total_curves * 0.58))
    valid_target = max(1, min(valid_target, total_curves - 1))
    invalid_target = total_curves - valid_target
    min_valid = min(MIN_PER_SUBTYPE, max(0, valid_target // max(1, len(VALID_SUBTYPES))))
    min_invalid = min(MIN_PER_SUBTYPE, max(0, invalid_target // max(1, len(INVALID_SUBTYPES))))

    valid_weights = {
        "plain": max(0.1, 1.0 - (pct_fast_growth + pct_late_growth + pct_valid_decline + 0.2)),
        "fast": pct_fast_growth,
        "late": pct_late_growth,
        "diauxic": 0.2,
        "decline": pct_valid_decline,
    }
    invalid_weights = {
        "nearreal": pct_nearreal_invalid_curves,
        "subtle": pct_subtle_invalid_curves,
        "obvious": pct_obvious_invalid_curves,
        "decline_only": pct_invalid_decline_only,
        "oscillation": pct_invalid_oscillation,
        "noise": pct_invalid_noise,
        "collapse": pct_invalid_collapse,
    }
    valid_counts = _allocate_counts(valid_target, valid_weights, min_valid)
    invalid_counts = _allocate_counts(invalid_target, invalid_weights, min_invalid)

    plan = [("VALID", k) for k, n in valid_counts.items() for _ in range(n)]
    plan += [("INVALID", k) for k, n in invalid_counts.items() for _ in range(n)]
    rng.shuffle(plan)

    bucket_tag = f"{bucket_h:g}h_{bucket_step:g}"
    rows = []
    for i, (cls, subtype) in enumerate(plan, start=1):
        is_valid = cls == "VALID"
        if is_valid:
            model_name, y = _sample_good_model_for_subtype(
                subtype=subtype,
                time_points=time_points,
                max_time=bucket_h,
                rng=rng,
            )
            if subtype == "decline":
                # Guard against invalid ranges on shorter horizons (e.g. 8h bucket).
                if bucket_h <= 16.0:
                    low = 8.0
                    high = min(12.0, bucket_h - bucket_step)
                else:
                    low = 10.0
                    high = min(14.0, bucket_h - bucket_step)
                if high <= low:
                    t_start = max(0.0, min(bucket_h * 0.75, bucket_h - bucket_step))
                else:
                    t_start = rng.uniform(low, high)
                y = apply_valid_decline(y, time_points, rng, t_start, rng.uniform(valid_decline_k_min, valid_decline_k_max))
            model_label = f"{model_name}_VALID_{subtype}"
            subtype_label = f"valid_{subtype}"
        else:
            base_model, base_curve = _sample_good_model_for_subtype(
                subtype="plain",
                time_points=time_points,
                max_time=bucket_h,
                rng=rng,
            )
            if subtype == "nearreal":
                y = make_near_real_invalid(base_curve, rng)
            elif subtype == "subtle":
                y = make_subtle_invalid(base_curve, rng)
            elif subtype == "obvious":
                y = make_obvious_invalid(base_curve, rng)
            elif subtype == "decline_only":
                y = make_decline_only_invalid(time_points, rng)
            elif subtype == "oscillation":
                y = make_chaotic_oscillation(time_points, rng)
            elif subtype == "noise":
                y = make_noise_dominated(time_points, rng) if rng.random() < 0.7 else flat_line(time_points, rng.uniform(0.0, 0.1))
            else:
                y = make_sudden_collapse(base_curve, rng)
            model_label = f"{base_model}_INVALID_{subtype}"
            subtype_label = f"invalid_{subtype}"

        noise_std = noise_level * 0.3 if (is_valid and rng.random() < pct_high_quality_valid) else noise_level
        y = (y + rng.normal(0, noise_std, size=y.shape)).clip(min=0.0)

        if is_valid:
            if rng.random() < pct_missing_curves:
                y = inject_missing(pd.Series(y), missing_frac_per_curve, rng).values
            if rng.random() < pct_outlier_curves:
                y = inject_negative_outliers(
                    pd.Series(y),
                    outlier_frac_per_curve,
                    outlier_scale_min,
                    outlier_scale_max,
                    rng,
                ).values

        row = {
            "FileName": f"{file_stem}_{bucket_tag}",
            "Test Id": f"{bucket_tag}_{i}",
            "Model Name": model_label,
            "Is_Valid": bool(is_valid),
            "Curve Subtype": subtype_label,
        }
        for j, t in enumerate(time_points):
            row[f"T{np.round(t, 6)} ({time_unit})"] = float(y[j])
        rows.append(row)

    return pd.DataFrame(rows), {
        "bucket": (bucket_h, bucket_step),
        "total": int(total_curves),
        "valid": int(valid_target),
        "invalid": int(invalid_target),
        "valid_counts": valid_counts,
        "invalid_counts": invalid_counts,
    }


def build_synthetic_training_set(
    *,
    train_data_dir: str | os.PathLike,
    out_syn_csv: str | os.PathLike,
    seed: int = 123,
    time_unit: str = "h",
) -> dict:
    lab_map = detect_lab_bucket_map(train_data_dir)
    bucket_summaries = []
    bucket_frames = []
    lab_inputs: list[str] = []
    subtype_count = len(VALID_SUBTYPES) + len(INVALID_SUBTYPES)

    for idx, (h, s) in enumerate(SUPPORTED_BUCKETS):
        labs = lab_map.get((h, s), [])
        for p in labs:
            if p not in lab_inputs:
                lab_inputs.append(p)
        n_lab = 0
        for p in labs:
            try:
                n_lab += int(len(pd.read_csv(p, low_memory=False)))
            except Exception:
                continue

        n_synth = max(
            MIN_SYNTHETIC_PER_BUCKET,
            SYNTHETIC_LAB_RATIO * n_lab,
            MIN_PER_SUBTYPE * subtype_count,
        )
        n_synth = int(min(MAX_SYNTHETIC_PER_BUCKET, n_synth))
        df_bucket, summary = generate_bucket_synthetic_df(
            bucket_h=h,
            bucket_step=s,
            total_curves=n_synth,
            file_stem="syn",
            seed=seed + idx * 97,
            time_unit=time_unit,
        )
        summary["lab_count"] = int(n_lab)
        summary["synth_count"] = int(n_synth)
        bucket_summaries.append(summary)
        bucket_frames.append(df_bucket)

    syn_df = pd.concat(bucket_frames, ignore_index=True) if bucket_frames else pd.DataFrame()
    out_syn_csv = os.fspath(out_syn_csv)
    os.makedirs(os.path.dirname(out_syn_csv), exist_ok=True)
    syn_df.to_csv(out_syn_csv, index=False)

    for sm in bucket_summaries:
        h, s = sm["bucket"]
        logging.info(
            "Bucket %sh step=%s: total=%d valid=%d invalid=%d lab=%d synth=%d",
            h, s, sm["total"], sm["valid"], sm["invalid"], sm["lab_count"], sm["synth_count"],
        )
        logging.info("  valid subtype counts: %s", sm["valid_counts"])
        logging.info("  invalid subtype counts: %s", sm["invalid_counts"])

    return {
        "syn_path": out_syn_csv,
        "lab_inputs": lab_inputs,
        "bucket_summaries": bucket_summaries,
        "total_synth": int(len(syn_df)),
    }
# --------------------------
# 3) Run-info writer (kept identical in spirit)
# --------------------------
def write_run_info_xlsx(output_dir, file_stem, args, wide_path, stats):
    """
    Appends run metadata to run_info.xlsx without overwriting prior runs.

    Sheets:
      - RUNS   (cumulative log; one row per run; appended)
      - INFO   (latest run snapshot; refreshed each execution)
      - PARAMS (latest run args;   refreshed each execution)
    """
    import os
    from datetime import datetime
    from openpyxl import Workbook, load_workbook

    xlsx_path = os.path.join(output_dir, "run_info.xlsx")
    os.makedirs(output_dir, exist_ok=True)

    # 1) Open or create workbook
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    # 2) Ensure RUNS sheet exists with header
    if "RUNS" in wb.sheetnames:
        ws_runs = wb["RUNS"]
        # If the sheet exists but is empty, write header
        if ws_runs.max_row == 1 and ws_runs.max_column == 1 and ws_runs["A1"].value is None:
            ws_runs.append([
                "Timestamp", "Output Dir", "Wide CSV", "Seed",
                "points_per_curve", "n_curves", "n_rows",
                "file_stem", "time_unit", "max_time", "time_step",
                "n_reps", "noise_level",
                "pct_missing_curves", "missing_frac_per_curve",
                "pct_outlier_curves", "outlier_frac_per_curve",
                "outlier_scale_min", "outlier_scale_max"
            ])
    else:
        ws_runs = wb.create_sheet("RUNS")
        ws_runs.append([
            "Timestamp", "Output Dir", "Wide CSV", "Seed",
            "points_per_curve", "n_curves", "n_rows",
            "file_stem", "time_unit", "max_time", "time_step",
            "n_reps", "noise_level",
            "pct_missing_curves", "missing_frac_per_curve",
            "pct_outlier_curves", "outlier_frac_per_curve",
            "outlier_scale_min", "outlier_scale_max"
        ])

    # 3) Append current run to RUNS
    ts = datetime.now().isoformat(timespec="seconds")
    ws_runs.append([
        ts,
        os.path.abspath(output_dir),
        os.path.abspath(wide_path),
        args.seed,
        stats.get("points_per_curve", 0),
        stats.get("n_curves", 0),
        stats.get("n_rows", 0),
        file_stem,
        getattr(args, "time_unit", None),
        getattr(args, "max_time", None),
        getattr(args, "time_step", None),
        getattr(args, "n_reps", None),
        getattr(args, "noise_level", None),
        getattr(args, "pct_missing_curves", None),
        getattr(args, "missing_frac_per_curve", None),
        getattr(args, "pct_outlier_curves", None),
        getattr(args, "outlier_frac_per_curve", None),
        getattr(args, "outlier_scale_min", None),
        getattr(args, "outlier_scale_max", None),
    ])

    # 4) Refresh INFO (latest snapshot)
    if "INFO" in wb.sheetnames:
        del wb["INFO"]
    ws_info = wb.create_sheet("INFO")
    a1_text = (
        f"Output: {os.path.abspath(output_dir)} | "
        f"File: {os.path.basename(wide_path)} | "
        f"Seed: {args.seed} | "
        f"Timestamp: {ts}"
    )
    ws_info["A1"] = a1_text
    ws_info["A3"] = "points_per_curve"; ws_info["B3"] = stats.get("points_per_curve", 0)
    ws_info["A4"] = "n_curves";         ws_info["B4"] = stats.get("n_curves", 0)
    ws_info["A5"] = "n_rows";           ws_info["B5"] = stats.get("n_rows", 0)

    # 5) Refresh PARAMS (latest args)
    if "PARAMS" in wb.sheetnames:
        del wb["PARAMS"]
    ws_params = wb.create_sheet("PARAMS")
    ws_params.append(["arg", "value"])
    for k, v in sorted(vars(args).items()):
        ws_params.append([k, str(v)])

    # 6) Save (in-place; preserves all old content and appended RUNS)
    # Remove the default 'Sheet' if it's still there and empty
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
        del wb["Sheet"]
    wb.save(xlsx_path)

# --------------------------
# 4) Main
# --------------------------
def main():
    p = argparse.ArgumentParser(description="Synthetic Growth Curve Generator (WIDE CSV only)")
    p.add_argument("--seed",                  type=int,   default=123)
    p.add_argument("--n-reps",                type=int,   default=10, help="replicates per model")
    p.add_argument("--max-time",              type=float, default=24.0, help="max time")
    p.add_argument("--time-step",             type=float, default=0.5, help="time step")
    p.add_argument("--time-unit",             type=str,   default="h", choices=["s","m","h"], help="unit label for header")
    p.add_argument("--noise-level",           type=float, default=0.05, help="Gaussian noise stdev")
    p.add_argument("--pct-missing-curves",    type=float, default=0.1,  help="fraction of curves with missing values")
    p.add_argument("--missing-frac-per-curve",type=float, default=0.1,  help="fraction of points set to NaN in selected curves")
    p.add_argument("--pct-outlier-curves",    type=float, default=0.05, help="fraction of curves with negative outliers")
    p.add_argument("--outlier-frac-per-curve",type=float, default=0.05, help="fraction of points made outliers in selected curves")
    p.add_argument("--outlier-scale-min",     type=float, default=0.1)
    p.add_argument("--outlier-scale-max",     type=float, default=0.3)
    p.add_argument("--output-dir",            type=str,   default="./dataNew")
    p.add_argument("--file-stem",             type=str,   default="timedata")
    # NEW: curve-quality and invalid patterns
    p.add_argument(
        "--pct-high-quality-valid",
        type=float,
        default=0.3,
        help="Fraction of GOOD model curves made very realistic (low noise, no missing/outliers)."
    )
    p.add_argument(
        "--pct-obvious-invalid-curves",
        type=float,
        default=0.1,
        help="Fraction of GOOD model curves corrupted into *obviously* invalid shapes."
    )
    p.add_argument(
        "--pct-subtle-invalid-curves",
        type=float,
        default=0.1,
        help="Fraction of GOOD model curves corrupted into *subtle* invalid shapes."
    )
    p.add_argument(
        "--pct-nearreal-invalid-curves",
        type=float,
        default=0.05,
        help="Fraction of GOOD model curves corrupted into almost-realistic but invalid shapes."
    )
    p.add_argument(
        "--pct-valid-decline",
        type=float,
        default=0.15,
        help="Fraction of GOOD model curves that are VALID with growth->peak->decline."
    )
    p.add_argument(
        "--pct-fast-growth",
        type=float,
        default=0.15,
        help="Fraction of GOOD model curves forced into fast-growth bucket."
    )
    p.add_argument(
        "--pct-late-growth",
        type=float,
        default=0.15,
        help="Fraction of GOOD model curves forced into late-growth bucket."
    )
    p.add_argument(
        "--valid-decline-start-min",
        type=float,
        default=None,
        help="If set, minimum time for decline start (overrides defaults based on max-time)."
    )
    p.add_argument(
        "--valid-decline-start-max",
        type=float,
        default=None,
        help="If set, maximum time for decline start (overrides defaults based on max-time)."
    )
    p.add_argument(
        "--valid-decline-k-min",
        type=float,
        default=0.03,
        help="Minimum decay rate k for valid decline."
    )
    p.add_argument(
        "--valid-decline-k-max",
        type=float,
        default=0.15,
        help="Maximum decay rate k for valid decline."
    )
    p.add_argument(
        "--pct-invalid-decline-only",
        type=float,
        default=0.08,
        help="Fraction of GOOD model curves corrupted into decline-without-rise invalids."
    )
    p.add_argument(
        "--pct-invalid-oscillation",
        type=float,
        default=0.06,
        help="Fraction of GOOD model curves corrupted into chaotic oscillations."
    )
    p.add_argument(
        "--pct-invalid-noise",
        type=float,
        default=0.06,
        help="Fraction of GOOD model curves corrupted into noise-dominated curves."
    )
    p.add_argument(
        "--pct-invalid-collapse",
        type=float,
        default=0.08,
        help="Fraction of GOOD model curves corrupted into sudden collapse curves."
    )

    args = p.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
    rng = np.random.default_rng(args.seed)

    # time grid
    time_points = np.arange(0, args.max_time + args.time_step/2, args.time_step)
    n_t = len(time_points)

    rows = []
    curve_id = 0
    test_id_prefix = args.file_stem[-2:]
    total_curves = int(len(MODEL_SPECS) * args.n_reps)
    valid_target = int(round(total_curves * 0.58))
    valid_target = max(1, min(valid_target, total_curves - 1))
    invalid_target = total_curves - valid_target

    min_valid = min(int(round(valid_target * 0.1)), 20)
    min_invalid = min(int(round(invalid_target * 0.1)), 20)

    valid_weights = {
        "plain": max(0.1, 1.0 - (args.pct_fast_growth + args.pct_late_growth + args.pct_valid_decline + 0.2)),
        "fast": args.pct_fast_growth,
        "late": args.pct_late_growth,
        "diauxic": 0.2,
        "decline": args.pct_valid_decline,
    }
    invalid_weights = {
        "obvious": args.pct_obvious_invalid_curves,
        "subtle": args.pct_subtle_invalid_curves,
        "nearreal": args.pct_nearreal_invalid_curves,
        "decline_only": args.pct_invalid_decline_only,
        "oscillation": args.pct_invalid_oscillation,
        "noise": args.pct_invalid_noise * 0.5,
        "no_growth": args.pct_invalid_noise * 0.5,
        "collapse": args.pct_invalid_collapse,
    }

    valid_counts = _allocate_counts(valid_target, valid_weights, min_valid)
    invalid_counts = _allocate_counts(invalid_target, invalid_weights, min_invalid)

    plan = [("VALID", k) for k, n in valid_counts.items() for _ in range(n)]
    plan += [("INVALID", k) for k, n in invalid_counts.items() for _ in range(n)]
    rng.shuffle(plan)

    def _sample_good_model(subtype: str):
        model_name = "Diauxic" if subtype == "diauxic" else rng.choice(sorted(GOOD_MODELS))
        if model_name == "Diauxic":
            pars = {
                "A1": rng.uniform(0.3, 1.0),
                "mu1": rng.uniform(0.2, 1.2),
                "lam1": rng.uniform(0.0, 5.0),
                "A2": rng.uniform(0.3, 1.0),
                "mu2": rng.uniform(0.2, 1.2),
                "lam2": rng.uniform(5.0, 15.0),
            }
            y = diauxic(time_points, **pars)
            return model_name, y

        if args.max_time <= 16.0:
            late_lam_range = (8.0, 12.0)
        else:
            late_lam_range = (10.0, 16.0)

        if subtype == "fast":
            mu_range = (0.9, 1.5)
            lam_range = (0.0, 2.0)
        elif subtype == "late":
            mu_range = (0.8, 1.5)
            lam_range = late_lam_range
        else:
            mu_range = (0.2, 1.5)
            lam_range = (0.0, 10.0)

        pars = {"A": rng.uniform(0.5, 2.0), "mu": rng.uniform(*mu_range), "lam": rng.uniform(*lam_range)}
        if model_name == "ModifiedGompertz":
            pars.update({"alpha": rng.uniform(0.0, 0.3), "tshift": rng.uniform(5.0, 12.0)})
        if model_name == "Richards":
            pars.update({"nu": rng.uniform(0.5, 2.0)})
        y = MODEL_SPECS[model_name][0](time_points, **pars)
        return model_name, y

    for cls, subtype in plan:
        curve_id += 1
        is_valid = cls == "VALID"
        corruption_tag = None

        if cls == "VALID":
            model_name, y = _sample_good_model(subtype)
            if subtype == "decline":
                if args.valid_decline_start_min is None or args.valid_decline_start_max is None:
                    if args.max_time <= 16.0:
                        t_start = rng.uniform(8.0, 12.0)
                    else:
                        t_start = rng.uniform(10.0, 14.0)
                else:
                    t_start = rng.uniform(args.valid_decline_start_min, args.valid_decline_start_max)
                k = rng.uniform(args.valid_decline_k_min, args.valid_decline_k_max)
                y = apply_valid_decline(y, time_points, rng, t_start, k)
        else:
            if subtype in {"obvious", "subtle", "nearreal", "collapse"}:
                model_name, y = _sample_good_model("plain")
                if subtype == "obvious":
                    y = make_obvious_invalid(y, rng); corruption_tag = "Invalid_Obvious"
                elif subtype == "subtle":
                    y = make_subtle_invalid(y, rng); corruption_tag = "Invalid_Subtle"
                elif subtype == "nearreal":
                    y = make_near_real_invalid(y, rng); corruption_tag = "Invalid_NearReal"
                else:
                    y = make_sudden_collapse(y, rng); corruption_tag = "Invalid_Collapse"
            elif subtype == "decline_only":
                model_name = "Flat"
                y = make_decline_only_invalid(time_points, rng); corruption_tag = "Invalid_DeclineOnly"
            elif subtype == "oscillation":
                model_name = "Flat"
                y = make_chaotic_oscillation(time_points, rng); corruption_tag = "Invalid_Oscillatory"
            elif subtype == "noise":
                model_name = "Flat"
                y = make_noise_dominated(time_points, rng); corruption_tag = "Invalid_Noise"
            else:
                model_name = "Flat"
                y = flat_line(time_points, rng.uniform(0.0, 0.1))
                corruption_tag = "Invalid_NoGrowth"

        # noise / clamp
        noise_std = args.noise_level * 0.3 if (is_valid and rng.random() < args.pct_high_quality_valid) else args.noise_level
        y = (y + rng.normal(0, noise_std, size=y.shape)).clip(min=0.0)

        # optional missing/outlier corruption (keep existing knobs; do not stack on invalids)
        if is_valid:
            if rng.random() < args.pct_missing_curves:
                y = inject_missing(pd.Series(y), args.missing_frac_per_curve, rng).values
            if rng.random() < args.pct_outlier_curves:
                y = inject_negative_outliers(
                    pd.Series(y), args.outlier_frac_per_curve,
                    args.outlier_scale_min, args.outlier_scale_max, rng
                ).values

        model_label = model_name
        if cls == "VALID" and subtype:
            model_label = f"{model_label}_VALID_{subtype}"
        if corruption_tag is not None:
            model_label = f"{model_label}_{corruption_tag}"

        base = {
            "FileName": args.file_stem,
            "Test Id": f"{test_id_prefix}_{curve_id}",
            "Model Name": model_label,
            "Is_Valid": bool(is_valid),
            "Curve Subtype": f"{str(cls).lower()}_{subtype}" if subtype else "",
        }

        for i, t in enumerate(time_points, start=1):
            base[f"T{np.round(t, 6)} ({args.time_unit})"] = float(y[i-1])
        rows.append(base)

    df_wide = pd.DataFrame(rows)

    # output
    os.makedirs(args.output_dir, exist_ok=True)
    wide_path = os.path.join(args.output_dir, f"timeseries_wide_{args.file_stem}.csv")
    df_wide.to_csv(wide_path, index=False)

    logging.info(
        f"Class balance: valid={valid_target} ({valid_target/total_curves:.2%}), "
        f"invalid={invalid_target} ({invalid_target/total_curves:.2%})"
    )
    logging.info(f"Valid subtype counts: {valid_counts}")
    logging.info(f"Invalid subtype counts: {invalid_counts}")

    # run-info stats (akin to long form, but derived from wide)
    stats = {
        "points_per_curve": int(n_t),
        "n_curves": int(len(df_wide)),
        "n_rows": int(len(df_wide)),  # one row per curve
    }
    write_run_info_xlsx(args.output_dir, args.file_stem, args, wide_path, stats)

    logging.info(f"Wrote wide CSV to {wide_path}")
    logging.info(f"Wrote run_info.xlsx to {os.path.join(args.output_dir, 'run_info.xlsx')}")

if __name__ == "__main__":
    main()
